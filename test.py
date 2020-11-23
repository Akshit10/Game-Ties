from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import requests
from bs4 import BeautifulSoup
import openpyxl
import time
from selenium.webdriver.common.by import By


### Locator Decalaration ###
gameList=(By.XPATH,"//div[@id='game_list']")
lenParent=(By.XPATH,"//section[3]/div[1]/div[1]/ul[1]")


### Main Function###
def getData():
    driver=openBrowser()
    listOfGames=getListofGames(driver)
    gamedata=getGamesNameandURL(driver,listOfGames)
    outputExcel(gamedata)

### Open Browser ###
def openBrowser():
    #Open Chrome Browser which is headless
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(),options=options)
    driver.get("https://www.game.tv/")
    print("Title of Page = " + driver.title)
    return driver

### List of Tiles ##
def getListofGames(driver):
    tournamentList = driver.find_element(*gameList)
    action = ActionChains(driver)
    action.move_to_element(tournamentList).perform()
    time.sleep(3)
    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    parentElement = driver.find_element(*lenParent)
    elementList = parentElement.find_elements_by_tag_name("li")
    print("Length of Games Tiles = {}".format(len(elementList)))
    return elementList

### Games Names and Url ###
def getGamesNameandURL(driver,listOfGames):

    gamesData={}
    for i in listOfGames:
        p_links = i.find_element_by_tag_name("a")
        name = p_links.find_element_by_tag_name("figcaption").text
        href = p_links.get_attribute('href')
        gamesData[name] = href
    print(gamesData)
    return gamesData

### Write Output in Excel ###
def outputExcel(gamedata):
    rows=1
    workopen_Games = openpyxl.load_workbook('Games Data.xlsx')
    sheet = workopen_Games.get_sheet_by_name('Data')
    for item, value in gamedata.items():
        rows += 1
        columns = 1
        sheet.cell(row=rows, column=columns).value = rows-1
        print("Game Name = " + item)
        columns += 1
        sheet.cell(row=rows, column=columns).value = item
        print("Tournament URL = " + value)
        columns += 1
        sheet.cell(row=rows, column=columns).value = value
        r = requests.get(value)
        print("Status Code = {} ".format(r.status_code))
        columns += 1
        sheet.cell(row=rows, column=columns).value = r.status_code
        html_content = r.text
        soup = BeautifulSoup(html_content)
        tornamentCount = soup.find(name='span', class_='count-tournaments').text
        columns += 1
        sheet.cell(row=rows, column=columns).value = tornamentCount
        print("Tournament Count = {}".format(tornamentCount))
        workopen_Games.save("Games Data.xlsx")

getData()